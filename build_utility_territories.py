#!/usr/bin/env python3
r"""
build_utility_territories.py
----------------------------
Resolves electric-utility EDO service territories to county FIPS and writes them into
edo_master_table_dual.json as `territory_geoids`.

WHY: all 31 Utility EDOs were built with territory_basis='home_county_only_INCOMPLETE' and a
territory of exactly ONE county (their HQ). That is a placeholder, not a fact -- so utilities only
ever surfaced as the serving EDO for their headquarters county (29 of 1518 counties total). Leads
in the rest of their real service area routed to nobody, or to a broader agency.

SOURCE: EIA Form EIA-861, "Service Territory" file -- the counties, by state, where each electric
utility has distribution equipment. Public, annual, authoritative.
    https://www.eia.gov/electricity/data/eia861/   ->  zip/f861<YEAR>.zip  ->  Service_Territory_<YEAR>.xlsx

USAGE:
    python build_utility_territories.py [path/to/Service_Territory_2024.xlsx]

Then ALWAYS rerun:  python build_edo_indexes.py     (routing indexes are derived from the master)

NOTE ON NAMES: FastLocations tracks utilities by their public/brand name ("Alliant Energy",
"AEP", "National Grid"); EIA reports the regulated operating subsidiary ("Interstate Power and
Light Co", "Ohio Power Co", "Niagara Mohawk Power Corp."). EIA_MAP below is that translation and is
the only part a human must maintain. Everything else is deterministic.
"""
import json, os, re, sys, unicodedata, math, collections

HERE = os.path.dirname(os.path.abspath(__file__))
MASTER = os.path.join(HERE, "edo_master_table_dual.json")
FEATURES = os.path.join(HERE, "county_features.json")
DEFAULT_XLSX = os.path.join(HERE, "Service_Territory_2024.xlsx")

# our objectid -> [(EIA utility name, restrict_to_state or None), ...]
EIA_MAP = {
    "4332": [("Ohio Power Co", None)],                      # AEP Ohio
    "41":   [("Alabama Power Co", None)],
    "4429": [("Interstate Power and Light Co", None)],      # Alliant Energy (IA)
    "4538": [("Wisconsin Power & Light Co", None)],         # Alliant Energy (WI)
    "4355": [("DTE Electric Company", None)],
    "4500": [("Entergy Louisiana LLC", None)],
    "4503": [("Entergy Texas Inc.", None)],
    "117":  [("Evergy Metro", "MO"), ("Evergy Missouri West", None)],
    "4537": [("Evergy Metro", "KS"), ("Evergy Kansas Central, Inc", None),
             ("Evergy Kansas South, Inc", None)],
    "14":   [("Florida Power & Light Co", None)],
    "4489": [("Georgia Power Co", None)],
    "4400": [("Greenville Utilities Comm", None)],          # genuinely single-county
    "4378": [("Indiana Michigan Power Co", None)],
    "89":   [("Kentucky Power Co", None)],
    "16":   [("Sierra Pacific Power Co", None)],            # NV Energy - Reno (northern NV)
    "4539": [("Nevada Power Co", None)],                    # NV Energy - Las Vegas (southern NV)
    "3":    [("Niagara Mohawk Power Corp.", None)],         # National Grid - Syracuse  }
    "4534": [("Niagara Mohawk Power Corp.", None)],         # National Grid - Buffalo   } split below
    "4535": [("Niagara Mohawk Power Corp.", None)],         # National Grid - Albany    }
    "4536": [("Long Island Power Authority", None)],        # National Grid - Long Island (see caveat)
    "4385": [("Rappahannock Electric Coop", None)],
    "15":   [("Salt River Project", None)],
}

# Offices that share one EIA utility's territory get it SPLIT: each county goes to whichever
# office is physically closest. Without this the three upstate NY offices would each claim all
# 37 Niagara Mohawk counties and compete for the same leads.
SPLIT_GROUPS = [["3", "4534", "4535"]]      # National Grid Syracuse / Buffalo / Albany

# Territory genuinely not derivable from EIA-861. Left as-is (home county, flagged INCOMPLETE) so
# the scorer's scope guard keeps discounting them rather than inventing coverage.
NOT_IN_EIA = {
    "4509": "G&T co-op (wholesale) - no retail territory reported to EIA",
    "4482": "G&T co-op (wholesale) - no retail territory reported to EIA",
    "4494": "wholesale river authority - no retail territory reported to EIA",
    "4423": "joint action agency - territory is its member municipalities",
    "101":  "joint action agency - territory is its member municipalities",
    "4333": "statewide co-op association - members not a single EIA entity",
    "100":  "marketing arm of the SC electric co-ops - not an EIA utility",
    "4569": "railroad, not an electric utility",
    "4432": "railroad, not an electric utility",
}


def cnorm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\b(county|parish|borough|census area|city and borough|municipality)\b", "", s)
    return re.sub(r"[^a-z0-9]", "", s)


def haversine_mi(a, b, c, d):
    R = 3958.8
    p1, p2 = math.radians(a), math.radians(c)
    dp, dl = math.radians(c - a), math.radians(d - b)
    h = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


def load_eia(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl required:  pip install openpyxl --break-system-packages")
    wb = openpyxl.load_workbook(path, read_only=True)
    util = collections.defaultdict(set)
    for sheet in wb.sheetnames:                      # Counties_States + Counties_Territories
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        next(it, None)                               # header
        for row in it:
            if row and row[2] and row[4] and row[5]:
                util[str(row[2]).strip()].add((str(row[4]).strip(), str(row[5]).strip()))
    return util


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        sys.exit(f"EIA service-territory file not found: {xlsx}\n"
                 f"Download f861<YEAR>.zip from https://www.eia.gov/electricity/data/eia861/ "
                 f"and extract Service_Territory_<YEAR>.xlsx")
    util = load_eia(xlsx)
    feat = json.load(open(FEATURES, encoding="utf-8"))
    lookup = {(d["ST_ABBREV"], cnorm(d["NAME"])): f for f, d in feat.items()}
    master = json.load(open(MASTER, encoding="utf-8"))
    by_id = {r["objectid"]: r for r in master}

    # 1. resolve each mapped utility to a FIPS set
    resolved, unmatched_names = {}, []
    for oid, specs in EIA_MAP.items():
        fips = set()
        for name, only_state in specs:
            if name not in util:
                unmatched_names.append((oid, name)); continue
            for st, cty in util[name]:
                if only_state and st != only_state:
                    continue
                f = lookup.get((st, cnorm(cty)))
                if f:
                    fips.add(f)
                else:
                    unmatched_names.append((oid, f"county {cty}, {st}"))
        resolved[oid] = fips

    # 2. split shared territories by nearest office
    for group in SPLIT_GROUPS:
        group = [o for o in group if o in resolved]
        if len(group) < 2:
            continue
        shared = set()
        for o in group:
            shared |= resolved[o]
        offices = []
        for o in group:
            r = by_id[o]
            try:
                offices.append((o, float(r["latitude"]), float(r["longitude"])))
            except (TypeError, ValueError):
                pass
        if len(offices) < 2:
            continue
        for o in group:
            resolved[o] = set()
        for f in shared:
            d = feat.get(f) or {}
            lat, lon = d.get("lat"), d.get("lon")
            if lat is None or lon is None:
                resolved[offices[0][0]].add(f); continue
            nearest = min(offices, key=lambda t: haversine_mi(lat, lon, t[1], t[2]))
            resolved[nearest[0]].add(f)

    # 3. write into the master table
    changed = []
    for oid, fips in resolved.items():
        if not fips:
            continue
        r = by_id[oid]
        before = len(r.get("territory_geoids") or [])
        geo = sorted(fips)
        r["territory_geoids"] = geo
        r["territory_fips"] = geo
        r["territory_county_count"] = len(geo)
        r["territory_basis"] = "eia861_service_territory"      # clears the _INCOMPLETE flag
        r["territory_source"] = os.path.basename(xlsx)
        changed.append((oid, r["organization"], before, len(geo)))
    for oid, why in NOT_IN_EIA.items():
        if oid in by_id:
            by_id[oid]["territory_note"] = why

    json.dump(master, open(MASTER, "w", encoding="utf-8"), indent=1, ensure_ascii=False)

    changed.sort(key=lambda t: -t[3])
    print(f"Resolved {len(changed)} utility territories from {os.path.basename(xlsx)}\n")
    for oid, org, b, a in changed:
        print(f"  #{oid:>4} {org[:40]:<40} {b:>2} -> {a:>3} counties")
    print(f"\n  left as-is (not EIA distribution utilities): {len(NOT_IN_EIA)}")
    if unmatched_names:
        print(f"\n  WARNING {len(unmatched_names)} unmatched name(s):")
        for oid, n in unmatched_names[:15]:
            print(f"    #{oid}  {n}")
    print(f"\nWrote {MASTER}\n>>> now run:  python build_edo_indexes.py")


if __name__ == "__main__":
    main()
